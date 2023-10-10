import json
from openpyxl import Workbook

# Load the JSON data
with open('/home/ubuntu/service_info.json', 'r') as json_file:
    service_info = json.load(json_file)

# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active

# Write service names and statuses to the Excel sheet
sheet['A1'] = 'Service Name'
sheet['B1'] = 'Status'

row = 2
for service_name, service_status in service_info['ansible_facts']['ansible_services'].items():
    sheet.cell(row=row, column=1, value=service_name)
    sheet.cell(row=row, column=2, value=service_status['state'])
    row += 1

# Save the Excel workbook
workbook.save('/home/ubuntu/service_info.json.xlsx')
